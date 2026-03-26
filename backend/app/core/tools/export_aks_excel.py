"""Export AKS registry to Excel.

Generates an Excel file with the complete AKS registry (equipment + cross-references).
This is the Phase 2 export — Revit matching sheets come in Phase 3.
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _style_header_row(ws, max_col: int):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = "A2"


def _auto_width(ws, max_col: int, max_row: int):
    for col in range(1, max_col + 1):
        max_len = 0
        for row in range(1, min(max_row + 1, 100)):
            val = ws.cell(row=row, column=col).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 50)


def export_aks_registry_excel(
    registry_data: dict,
    output_path: str | Path,
    on_progress: callable = None,
) -> str:
    """Generate AKS registry Excel file.

    Args:
        registry_data: Ergebnis von build_registry()
        output_path: Pfad fuer die Excel-Datei
        on_progress: Callback(progress_pct, message)

    Returns:
        Pfad zur generierten Excel-Datei
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    if on_progress:
        on_progress(10, "Erstelle AKS-Komplett Sheet...")

    # Sheet 1: AKS Komplett
    ws = wb.active
    ws.title = "AKS Komplett"

    headers = [
        "AKS", "Raum", "Anlage", "Gewerk", "Geraet",
        "Geraet-Typ", "Hat Schema", "Schema-Kinder",
        "Schema-Seiten", "PDF X", "PDF Y",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    row = 2
    for eq in registry_data.get("equipment", []):
        ws.cell(row=row, column=1, value=eq.get("aks_parent"))
        ws.cell(row=row, column=2, value=eq.get("room"))
        ws.cell(row=row, column=3, value=eq.get("anlage"))
        ws.cell(row=row, column=4, value=eq.get("gewerk"))
        ws.cell(row=row, column=5, value=eq.get("geraet"))
        ws.cell(row=row, column=6, value=eq.get("geraet_type"))
        ws.cell(row=row, column=7, value="Ja" if eq.get("has_schema") else "Nein")
        ws.cell(row=row, column=8, value=len(eq.get("schema_children", [])))
        ws.cell(row=row, column=9, value=", ".join(str(p) for p in eq.get("schema_pages", [])))
        ws.cell(row=row, column=10, value=eq.get("pdf_x"))
        ws.cell(row=row, column=11, value=eq.get("pdf_y"))

        fill = GREEN_FILL if eq.get("has_schema") else RED_FILL
        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).fill = fill
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    _style_header_row(ws, len(headers))
    _auto_width(ws, len(headers), row)

    if on_progress:
        on_progress(50, "Erstelle Querverweise Sheet...")

    # Sheet 2: Querverweise
    ws2 = wb.create_sheet("Querverweise")
    headers2 = [
        "Text", "Typ", "Anlage", "Beschreibung",
        "Raum-Ref", "Schema-Seiten", "Aufgeloest", "PDF X", "PDF Y",
    ]
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h)

    row = 2
    for ref in registry_data.get("cross_references", []):
        ws2.cell(row=row, column=1, value=ref.get("text"))
        ws2.cell(row=row, column=2, value=ref.get("type"))
        ws2.cell(row=row, column=3, value=ref.get("anlage"))
        ws2.cell(row=row, column=4, value=ref.get("description"))
        ws2.cell(row=row, column=5, value=ref.get("room_ref"))
        pages = ref.get("schema_pages", [])
        ws2.cell(row=row, column=6, value=", ".join(str(p) for p in pages) if pages else "")
        ws2.cell(row=row, column=7, value="Ja" if ref.get("resolved") else "Nein")
        ws2.cell(row=row, column=8, value=ref.get("pdf_x"))
        ws2.cell(row=row, column=9, value=ref.get("pdf_y"))

        fill = GREEN_FILL if ref.get("resolved") else RED_FILL
        for c in range(1, len(headers2) + 1):
            ws2.cell(row=row, column=c).fill = fill
        row += 1

    _style_header_row(ws2, len(headers2))
    _auto_width(ws2, len(headers2), row)

    if on_progress:
        on_progress(80, "Erstelle Raum-Uebersicht Sheet...")

    # Sheet 3: Raum-Uebersicht
    ws3 = wb.create_sheet("Raum-Uebersicht")
    headers3 = ["Raum", "Anzahl AKS", "Geraet-Typen"]
    for col, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=col, value=h)

    row = 2
    for room, aks_list in sorted(registry_data.get("room_index", {}).items()):
        types: dict[str, int] = {}
        for eq in registry_data.get("equipment", []):
            if eq["room"] == room:
                t = eq.get("geraet_type", "?")
                types[t] = types.get(t, 0) + 1
        type_str = ", ".join(f"{t}: {c}" for t, c in sorted(types.items()))

        ws3.cell(row=row, column=1, value=room)
        ws3.cell(row=row, column=2, value=len(aks_list))
        ws3.cell(row=row, column=3, value=type_str)
        row += 1

    _style_header_row(ws3, len(headers3))
    _auto_width(ws3, len(headers3), row)

    wb.save(str(output_path))
    wb.close()

    if on_progress:
        on_progress(100, "Excel erstellt")

    return str(output_path)
