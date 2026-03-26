"""Export Revit-Import Excel with AKS assignments.

Generates the 4-sheet Excel file for Revit reimport and review.
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CONFIDENCE_FILLS = {"HIGH": GREEN_FILL, "MEDIUM": YELLOW_FILL, "LOW": RED_FILL}


def _style_header_row(ws, max_col: int):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = "A2"


def _auto_width(ws, max_col: int, max_row: int):
    for col in range(1, max_col + 1):
        max_len = 0
        for row in range(1, min(max_row + 1, 100)):
            val = ws.cell(row=row, column=col).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 50)


def _create_revit_aks_sheet(wb, match_data: dict, original_excel_path: str | Path):
    """Sheet 1: Original Revit-Daten + AKS-Spalte (reimport-fertig)."""
    ws = wb.active
    ws.title = "Revit + AKS"

    orig_wb = openpyxl.load_workbook(str(original_excel_path))
    orig_ws = orig_wb.active
    orig_rows = list(orig_ws.iter_rows(values_only=True))

    if not orig_rows:
        orig_wb.close()
        return

    # Match-Lookup nach GUID
    match_by_guid = {m["revit_guid"]: m for m in match_data.get("matches", [])}

    headers = list(orig_rows[0]) + ["AKS", "AKS_Confidence"]
    for col, val in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=val)

    # GUID-Spalte finden
    guid_col = None
    for col_idx, h in enumerate(orig_rows[0]):
        if h and "ifcguid" in str(h).lower():
            guid_col = col_idx
            break

    for row_idx, row in enumerate(orig_rows[1:], 2):
        for col, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col, value=val)

        guid = row[guid_col] if guid_col is not None else None
        match = match_by_guid.get(guid)
        aks_col = len(headers) - 1
        conf_col = len(headers)

        if match:
            ws.cell(row=row_idx, column=aks_col, value=match["aks"])
            ws.cell(row=row_idx, column=conf_col, value=match["confidence"])
            fill = CONFIDENCE_FILLS.get(match["confidence"], GREEN_FILL)
            ws.cell(row=row_idx, column=aks_col).fill = fill
            ws.cell(row=row_idx, column=conf_col).fill = fill
        else:
            ws.cell(row=row_idx, column=aks_col, value="NICHT ZUGEORDNET")
            ws.cell(row=row_idx, column=conf_col, value="N/A")
            ws.cell(row=row_idx, column=aks_col).fill = RED_FILL

    _style_header_row(ws, len(headers))
    _auto_width(ws, len(headers), len(orig_rows))
    orig_wb.close()


def _create_matching_review_sheet(wb, match_data: dict):
    """Sheet 2: Zuordnungs-Tabelle zur Kontrolle."""
    ws = wb.create_sheet("Zuordnung")

    headers = [
        "Raum", "AKS", "Revit GUID", "Revit Typ",
        "X (Revit)", "Y (Revit)", "X (PDF)", "Y (PDF)",
        "Sortier-Achse", "Rang", "Konfidenz",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    row = 2
    for m in match_data.get("matches", []):
        ws.cell(row=row, column=1, value=m["room"])
        ws.cell(row=row, column=2, value=m["aks"])
        ws.cell(row=row, column=3, value=m["revit_guid"])
        ws.cell(row=row, column=4, value=m.get("revit_type", ""))
        ws.cell(row=row, column=5, value=m.get("revit_x"))
        ws.cell(row=row, column=6, value=m.get("revit_y"))
        ws.cell(row=row, column=7, value=m.get("pdf_x"))
        ws.cell(row=row, column=8, value=m.get("pdf_y"))
        ws.cell(row=row, column=9, value=m.get("sort_axis"))
        ws.cell(row=row, column=10, value=m.get("sort_rank"))
        ws.cell(row=row, column=11, value=m.get("confidence"))

        fill = CONFIDENCE_FILLS.get(m.get("confidence"), GREEN_FILL)
        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).fill = fill
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    for um in match_data.get("unmatched_aks", []):
        ws.cell(row=row, column=1, value=um.get("room"))
        ws.cell(row=row, column=2, value=um.get("aks"))
        ws.cell(row=row, column=11, value=f"UNMATCHED: {um.get('reason')}")
        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).fill = RED_FILL
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    _style_header_row(ws, len(headers))
    _auto_width(ws, len(headers), row)


def _create_room_summary_sheet(wb, match_data: dict):
    """Sheet 3: Raum-Zusammenfassung."""
    ws = wb.create_sheet("Raum-Uebersicht")

    headers = ["Raum", "Status", "AKS", "Revit", "Matched", "Methode", "Konfidenz"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    row = 2
    for room, summary in sorted(match_data.get("room_summary", {}).items()):
        ws.cell(row=row, column=1, value=room)
        ws.cell(row=row, column=2, value=summary["status"])
        ws.cell(row=row, column=3, value=summary.get("aks_count", 0))
        ws.cell(row=row, column=4, value=summary.get("revit_count", 0))
        ws.cell(row=row, column=5, value=summary.get("matched", 0))
        ws.cell(row=row, column=6, value=summary.get("method", ""))
        ws.cell(row=row, column=7, value=summary.get("confidence", ""))

        status = summary["status"]
        fill = GREEN_FILL if status == "MATCHED" else (YELLOW_FILL if "MISMATCH" in status else RED_FILL)
        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).fill = fill
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    _style_header_row(ws, len(headers))
    _auto_width(ws, len(headers), row)


def _create_unmatched_sheet(wb, match_data: dict):
    """Sheet 4: Nicht zugeordnete Elemente."""
    ws = wb.create_sheet("Nicht zugeordnet")

    headers = ["Typ", "Raum", "AKS / GUID", "Grund"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    row = 2
    for um in match_data.get("unmatched_aks", []):
        ws.cell(row=row, column=1, value="AKS")
        ws.cell(row=row, column=2, value=um.get("room"))
        ws.cell(row=row, column=3, value=um.get("aks"))
        ws.cell(row=row, column=4, value=um.get("reason"))
        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).fill = RED_FILL
        row += 1

    for um in match_data.get("unmatched_revit", []):
        ws.cell(row=row, column=1, value="Revit")
        ws.cell(row=row, column=2, value=um.get("room"))
        ws.cell(row=row, column=3, value=um.get("guid"))
        ws.cell(row=row, column=4, value=um.get("reason"))
        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).fill = YELLOW_FILL
        row += 1

    _style_header_row(ws, len(headers))
    _auto_width(ws, len(headers), row)


def export_revit_import_excel(
    match_data: dict,
    original_excel_path: str | Path,
    output_path: str | Path,
    on_progress: callable = None,
) -> str:
    """Revit-Import-Excel mit AKS-Zuordnungen generieren.

    Args:
        match_data: Ergebnis von match_revit_to_aks()
        original_excel_path: Pfad zur Original-Revit-Excel
        output_path: Pfad fuer die Ausgabe-Excel
        on_progress: Callback(progress_pct, message)

    Returns:
        Pfad zur generierten Excel-Datei
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    if on_progress:
        on_progress(10, "Erstelle Revit+AKS Sheet...")
    _create_revit_aks_sheet(wb, match_data, original_excel_path)

    if on_progress:
        on_progress(35, "Erstelle Zuordnungs-Sheet...")
    _create_matching_review_sheet(wb, match_data)

    if on_progress:
        on_progress(55, "Erstelle Raum-Uebersicht...")
    _create_room_summary_sheet(wb, match_data)

    if on_progress:
        on_progress(75, "Erstelle Unmatched-Sheet...")
    _create_unmatched_sheet(wb, match_data)

    wb.save(str(output_path))
    wb.close()

    if on_progress:
        on_progress(100, "Excel erstellt")

    return str(output_path)
